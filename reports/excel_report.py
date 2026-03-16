"""
Excel Report Generator for WebScanPro

This module provides functionality to generate Excel reports for security scan results.
It uses the openpyxl library to create professional-looking Excel workbooks.
"""

import os
import logging
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple

# Try to import openpyxl, but don't fail if it's not installed
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment, 
        NamedStyle, Protection, GradientFill, Color
    )
    from openpyxl.styles.colors import WHITE, BLACK
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
    from openpyxl.worksheet.page import PageMargins, PrintPageSetup
    from openpyxl.worksheet.views import SheetView
    from openpyxl.drawing.image import Image as XLImage
    
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Set up logger
logger = logging.getLogger('webscanpro.excel_report')

# Define styles
class ExcelStyles:
    """Container for Excel styles."""
    
    def __init__(self, workbook):
        """Initialize styles for the Excel workbook."""
        self.workbook = workbook
        self._create_styles()
    
    def _create_styles(self):
        """Create and register styles for the Excel workbook."""
        # Title style
        self.title_style = NamedStyle(name="title_style")
        self.title_style.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        self.title_style.fill = PatternFill("solid", fgColor="2C3E50")
        self.title_style.alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.add_named_style(self.title_style)
        
        # Header style
        self.header_style = NamedStyle(name="header_style")
        self.header_style.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        self.header_style.fill = PatternFill("solid", fgColor="3498DB")
        self.header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.workbook.add_named_style(self.header_style)
        
        # Normal style
        self.normal_style = NamedStyle(name="normal_style")
        self.normal_style.font = Font(name='Calibri', size=10)
        self.normal_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.normal_style.alignment = Alignment(vertical='top', wrap_text=True)
        self.workbook.add_named_style(self.normal_style)
        
        # Code style
        self.code_style = NamedStyle(name="code_style")
        self.code_style.font = Font(name='Consolas', size=9)
        self.code_style.fill = PatternFill("solid", fgColor="F8F9FA")
        self.code_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.code_style.alignment = Alignment(vertical='top', wrap_text=True)
        self.workbook.add_named_style(self.code_style)
        
        # Severity styles
        self.severity_styles = {
            'high': self._create_severity_style('high', 'DC3545', 'FFFFFF'),
            'medium': self._create_severity_style('medium', 'FFC107', '000000'),
            'low': self._create_severity_style('low', '17A2B8', 'FFFFFF'),
            'info': self._create_severity_style('info', '6C757D', 'FFFFFF')
        }
    
    def _create_severity_style(self, name: str, bg_color: str, font_color: str) -> NamedStyle:
        """Create a style for a specific severity level."""
        style = NamedStyle(name=f"severity_{name}_style")
        style.font = Font(name='Calibri', size=10, bold=True, color=font_color)
        style.fill = PatternFill("solid", fgColor=bg_color)
        style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        style.alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.add_named_style(style)
        return style

def _auto_adjust_columns(worksheet):
    """Auto-adjust column widths based on content."""
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        
        # Find the maximum length of content in the column
        for cell in column_cells:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Add some padding
        adjusted_width = (max_length + 2) * 1.1
        
        # Set a maximum width to prevent extremely wide columns
        max_width = 50
        adjusted_width = min(adjusted_width, max_width)
        
        # Set the column width
        worksheet.column_dimensions[column].width = adjusted_width

def _add_summary_sheet(workbook, scan_results, styles):
    """Add a summary sheet to the workbook."""
    ws = workbook.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    
    # Set print settings
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # Add title
    ws.merge_cells('A1:B1')
    ws['A1'] = "WebScanPro Security Report"
    ws['A1'].style = styles.title_style
    
    # Add scan information
    ws.merge_cells('A2:B2')
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True)
    
    ws.merge_cells('A3:B3')
    ws['A3'] = f"Target: {scan_results.get('target', 'N/A')}"
    ws['A3'].font = Font(name='Calibri', size=10, bold=True)
    
    # Add summary table
    summary_data = [
        ["Scan ID", scan_results.get('scan_id', 'N/A')],
        ["Start Time", scan_results.get('start_time', 'N/A')],
        ["End Time", scan_results.get('end_time', 'N/A')],
        ["Status", scan_results.get('status', 'N/A').capitalize()],
        ["Total Vulnerabilities", len(scan_results.get('vulnerabilities', []))]
    ]
    
    # Add vulnerability counts by severity
    vulnerabilities = scan_results.get('vulnerabilities', [])
    severity_counts = {
        'High': 0,
        'Medium': 0,
        'Low': 0,
        'Info': 0
    }
    
    for vuln in vulnerabilities:
        severity = vuln.get('severity', 'info').capitalize()
        if severity in severity_counts:
            severity_counts[severity] += 1
        else:
            severity_counts['Info'] += 1
    
    summary_data.extend([
        ["High Severity", severity_counts['High']],
        ["Medium Severity", severity_counts['Medium']],
        ["Low Severity", severity_counts['Low']],
        ["Informational", severity_counts['Info']]
    ])
    
    # Write summary data to the worksheet
    for i, (label, value) in enumerate(summary_data, start=5):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(name='Calibri', size=10, bold=True)
        
        ws[f'B{i}'] = value
        ws[f'B{i}'].font = Font(name='Calibri', size=10)
    
    # Add a table of contents for vulnerability details
    if vulnerabilities:
        ws['A15'] = "Vulnerability Details"
        ws['A15'].font = Font(name='Calibri', size=12, bold=True)
        
        for i, vuln in enumerate(vulnerabilities, start=16):
            vuln_id = vuln.get('id', f'VULN-{i-15}')
            vuln_type = vuln.get('type', 'Unknown')
            severity = vuln.get('severity', 'info').capitalize()
            
            ws[f'A{i}'] = f"{vuln_id}: {vuln_type}"
            ws[f'A{i}'].font = Font(name='Calibri', size=10, underline='single', color='0563C1')
            ws[f'A{i}'].hyperlink = f"'Vulnerability {i-15}'!A1"
            
            ws[f'B{i}'] = severity
            if severity.lower() in styles.severity_styles:
                ws[f'B{i}'].style = styles.severity_styles[severity.lower()].name
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    
    # Add a link back to the summary sheet in all vulnerability sheets
    workbook.create_named_range('_Summary', ws, 'A1')

def _add_vulnerability_sheets(workbook, scan_results, styles):
    """Add a sheet for each vulnerability to the workbook."""
    vulnerabilities = scan_results.get('vulnerabilities', [])
    
    for i, vuln in enumerate(vulnerabilities, start=1):
        # Create a new worksheet for the vulnerability
        ws = workbook.create_sheet(f"Vulnerability {i}")
        ws.sheet_view.showGridLines = False
        
        # Set print settings
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        # Add title
        vuln_id = vuln.get('id', f'VULN-{i}')
        vuln_type = vuln.get('type', 'Unknown')
        severity = vuln.get('severity', 'info').lower()
        
        ws.merge_cells('A1:B1')
        ws['A1'] = f"{vuln_id}: {vuln_type}"
        ws['A1'].style = styles.title_style
        
        # Add severity
        ws.merge_cells('A2:B2')
        ws['A2'] = f"Severity: {severity.upper()}"
        if severity in styles.severity_styles:
            ws['A2'].style = styles.severity_styles[severity].name
        else:
            ws['A2'].style = styles.severity_styles['info'].name
        
        # Add vulnerability details
        details = [
            ("Description", vuln.get('description', 'No description available.')),
            ("URL", vuln.get('url', 'N/A')),
            ("Parameter", vuln.get('parameter', 'N/A')),
            ("Impact", vuln.get('impact', 'N/A')),
            ("Remediation", vuln.get('remediation', 'No remediation information available.'))
        ]
        
        row = 4
        for label, value in details:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
            
            ws[f'B{row}'] = value
            ws[f'B{row}'].style = styles.normal_style
            
            row += 1
        
        # Add evidence if available
        if 'evidence' in vuln:
            evidence = vuln['evidence']
            
            ws[f'A{row}'] = "Evidence"
            ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
            row += 1
            
            # Add request
            ws[f'A{row}'] = "Request:"
            ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
            
            ws[f'B{row}'] = evidence.get('request', 'N/A')
            ws[f'B{row}'].style = styles.code_style
            row += 1
            
            # Add response code
            ws[f'A{row}'] = "Response Code:"
            ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
            
            ws[f'B{row}'] = str(evidence.get('response_code', 'N/A'))
            ws[f'B{row}'].style = styles.normal_style
            row += 1
            
            # Add response snippet
            ws[f'A{row}'] = "Response Snippet:"
            ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
            
            ws[f'B{row}'] = evidence.get('response_body_snippet', 'N/A')
            ws[f'B{row}'].style = styles.code_style
            row += 1
        
        # Add references if available
        if 'references' in vuln and vuln['references']:
            ws[f'A{row}'] = "References:"
            ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
            row += 1
            
            for ref in vuln['references']:
                ws[f'B{row}'] = ref
                ws[f'B{row}'].font = Font(name='Calibri', size=10, underline='single', color='0563C1')
                ws[f'B{row}'].hyperlink = ref
                row += 1
        
        # Add a link back to the summary
        ws[f'A{row}'] = "<< Back to Summary"
        ws[f'A{row}'].font = Font(name='Calibri', size=10, underline='single', color='0563C1')
        ws[f'A{row}'].hyperlink = "#_Summary"
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 80
        
        # Set row heights
        for r in range(1, row + 1):
            ws.row_dimensions[r].height = 15
        
        # Set column heights for multi-line cells
        ws.row_dimensions[4].height = 30  # Description
        ws.row_dimensions[8].height = 30  # Impact
        ws.row_dimensions[9].height = 30  # Remediation
        
        # Add page breaks between sections if needed
        # ws.row_breaks.append(Break(id=row-1))

def generate(scan_results: Dict[str, Any], output_path: str, **kwargs) -> None:
    """
    Generate an Excel report from scan results.
    
    Args:
        scan_results: Dictionary containing scan results
        output_path: Path to save the Excel report
        **kwargs: Additional keyword arguments (e.g., title, author, version)
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl is not installed. Please install it with 'pip install openpyxl'")
        raise ImportError("openpyxl is required for Excel report generation.")
    
    try:
        # Create a new workbook
        wb = Workbook()
        
        # Remove the default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Set document properties
        wb.properties.title = kwargs.get('title', 'WebScanPro Security Report')
        wb.properties.creator = kwargs.get('author', 'WebScanPro')
        wb.properties.version = kwargs.get('version', '1.0.0')
        
        # Create styles
        styles = ExcelStyles(wb)
        
        # Add summary sheet
        _add_summary_sheet(wb, scan_results, styles)
        
        # Add vulnerability sheets
        _add_vulnerability_sheets(wb, scan_results, styles)
        
        # Set the active sheet to the summary
        wb.active = wb['Summary']
        
        # Save the workbook
        wb.save(output_path)
        
        logger.info(f"Generated Excel report: {output_path}")
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}", exc_info=True)
        raise

if __name__ == "__main__":
    # Example usage
    example_results = {
        "scan_id": "scan_12345",
        "target": "https://example.com",
        "start_time": "2026-02-18T10:00:00+05:30",
        "end_time": "2026-02-18T10:30:00+05:30",
        "status": "completed",
        "vulnerabilities_found": 2,
        "vulnerabilities": [
            {
                "id": "vuln_001",
                "type": "SQL Injection",
                "severity": "High",
                "url": "https://example.com/search?q=test",
                "parameter": "q",
                "description": "SQL injection vulnerability found in the search parameter. The application is vulnerable to boolean-based blind SQL injection attacks.",
                "impact": "An attacker could extract sensitive information from the database, modify database queries, or potentially gain unauthorized access to the system.",
                "remediation": "Use parameterized queries or prepared statements to prevent SQL injection. Input validation and output encoding should also be implemented.",
                "evidence": {
                    "request": "GET /search?q=test' OR '1'='1 HTTP/1.1\nHost: example.com\nUser-Agent: Mozilla/5.0\nAccept: text/html,application/xhtml+xml\n\n",
                    "response_code": 500,
                    "response_body_snippet": "Error: You have an error in your SQL syntax; check the manual that corresponds to your MySQL server version for the right syntax to use near '1'' at line 1"
                },
                "references": [
                    "https://owasp.org/www-community/attacks/SQL_Injection",
                    "https://portswigger.net/web-security/sql-injection"
                ]
            },
            {
                "id": "vuln_002",
                "type": "XSS",
                "severity": "Medium",
                "url": "https://example.com/contact",
                "parameter": "message",
                "description": "Cross-site scripting (XSS) vulnerability found in the contact form. The application does not properly sanitize user input in the message parameter.",
                "impact": "An attacker could execute arbitrary JavaScript code in the context of the victim's browser, potentially leading to session hijacking, defacement, or redirection to malicious sites.",
                "remediation": "Implement proper input validation and output encoding. Use Content Security Policy (CSP) headers to mitigate the impact of XSS vulnerabilities.",
                "evidence": {
                    "request": "POST /contact HTTP/1.1\nHost: example.com\nContent-Type: application/x-www-form-urlencoded\n\nname=test&email=test@example.com&message=<script>alert('XSS')</script>",
                    "response_code": 200,
                    "response_body_snippet": "<div class=\"message\"><script>alert('XSS')</script></div>"
                },
                "references": [
                    "https://owasp.org/www-community/attacks/xss/",
                    "https://portswigger.net/web-security/cross-site-scripting"
                ]
            }
        ]
    }
    
    # Generate the report
    generate(example_results, "example_report.xlsx", title="Example Security Report", version="1.0.0")
